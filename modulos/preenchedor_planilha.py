"""
Preenchedor de planilha urbanistica — JSON do pipeline -> XLSX consolidado.

Fluxo:
  1. Recebe lista de processamento_id (JSONs do pipeline)
  2. Ordena por data_publicacao (mais antiga primeiro)
  3. Aplica "lei mais recente ganha" por (zona, parametro)
  4. Preenche planilha base (v3.2 ou outra)
  5. Salva em /static/planilhas_geradas/<slug>/

Regras de preenchimento:
  - Uso explicitamente listado, status SIM ou CONDICIONADO  -> SIM + parametros
  - Uso explicitamente listado, status NÃO                  -> NÃO, parametros vazios
  - Uso silenciado (nao aparece em NENHUMA zona da lei)     -> NI em tudo
  - Variacoes condicionais: multi-linha "valor_padrao\\nvalor_alternativo se cond"
"""
import os
import json
import logging
from datetime import datetime
from typing import Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────
# CONSTANTES — mapeamento JSON keys -> ordem na planilha v3.2
# ─────────────────────────────────────────────────────────────────────

# Usos na ordem da planilha v3.2 (9 usos)
USOS_PLANILHA = [
    'residencial_unifamiliar',
    'residencial_multifamiliar',
    'residencial_his',
    'residencial_transitorio_hotel',  # novo
    'comercial',
    'servicos',
    'uso_misto',
    'industrial',
    'institucional',
]

# Labels humanos (debug/log)
USOS_LABEL = {
    'residencial_unifamiliar': 'Residencial Unifamiliar',
    'residencial_multifamiliar': 'Residencial Multifamiliar',
    'residencial_his': 'Residencial HIS',
    'residencial_transitorio_hotel': 'Residencial Transitorio / Hotel',
    'comercial': 'Comercial',
    'servicos': 'Servicos',
    'uso_misto': 'Uso Misto',
    'industrial': 'Industrial',
    'institucional': 'Institucional',
}

# ─────────────────────────────────────────────────────────────────────
# 1. CARREGAR + ORDENAR JSONs
# ─────────────────────────────────────────────────────────────────────

def carregar_jsons(jsons_ids: List[int], get_db_func) -> List[Dict]:
    """
    Carrega JSONs do banco/disco pelos IDs em legislacao_processamentos.
    Retorna lista ordenada por data_publicacao ASC (mais antiga primeiro).
    """
    if not jsons_ids:
        return []
    import psycopg2.extras
    conn = get_db_func()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute(
        "SELECT id, output_dir, legislacao_label, processado_em "
        "FROM legislacao_processamentos WHERE id = ANY(%s) AND sucesso = TRUE",
        (jsons_ids,)
    )
    rows = cur.fetchall()
    cur.close(); conn.close()

    jsons = []
    for r in rows:
        out_dir = r.get('output_dir')
        if not out_dir:
            logger.warning(f"proc_id={r['id']} sem output_dir")
            continue
        json_path = os.path.join(out_dir, 'resultado_final.json')
        if not os.path.exists(json_path):
            logger.warning(f"proc_id={r['id']} sem resultado_final.json em {out_dir}")
            continue
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                d = json.load(f)
        except Exception as e:
            logger.error(f"proc_id={r['id']} falhou ao ler JSON: {e}")
            continue

        # Extrai data_publicacao
        data_pub = (d.get('estado', {}).get('legislacao', {}) or {}).get('data_publicacao') or ''
        jsons.append({
            'proc_id': r['id'],
            'legislacao_label': r['legislacao_label'],
            'data_publicacao': data_pub,
            'data': d,
        })

    # Ordena por data_publicacao ASC (mais antiga primeiro)
    # JSONs sem data ficam ao final (mas ainda assim aplicados antes do consolidate)
    jsons.sort(key=lambda j: j['data_publicacao'] or '0000-00-00')
    return jsons


def fonte_lei(legislacao_meta: Dict) -> str:
    """Retorna 'Lei Complementar 148/2023' a partir do meta da legislacao."""
    if not legislacao_meta:
        return ''
    tipo = legislacao_meta.get('tipo', 'Lei')
    num = legislacao_meta.get('numero', '?')
    ano = legislacao_meta.get('ano', '?')
    return f"{tipo} {num}/{ano}"


# ─────────────────────────────────────────────────────────────────────
# 2. DETERMINAR USOS RECONHECIDOS pelo município (todos os JSONs)
# ─────────────────────────────────────────────────────────────────────

def usos_reconhecidos(jsons: List[Dict]) -> set:
    """Coleta TODOS os usos que apareceram em alguma zona em qualquer JSON.
    Usos fora deste set sao tratados como NI (lei silenciou)."""
    reconhecidos = set()
    for j in jsons:
        zonas = j['data'].get('estado', {}).get('zonas', {}) or {}
        for z_dados in zonas.values():
            usos = z_dados.get('usos_permitidos', {}) or {}
            reconhecidos.update(usos.keys())
    return reconhecidos


# ─────────────────────────────────────────────────────────────────────
# 3. CONSOLIDAR — aplica "lei mais recente ganha"
# ─────────────────────────────────────────────────────────────────────

def consolidar(jsons: List[Dict]) -> Dict:
    """
    Itera JSONs em ordem cronologica (antiga -> recente).
    Sobrescreve campos quando lei mais recente os redefine.
    Retorna {zonas: {zona_nome: {dados_completos + fontes}}}
    """
    consolidado = {'zonas': {}, 'fonte_geral': {}, 'fontes_modificacoes': []}

    for j in jsons:
        leg_meta = j['data'].get('estado', {}).get('legislacao', {}) or {}
        fonte = fonte_lei(leg_meta)

        # 1) Hierarquia viaria (top-level)
        hv = j['data'].get('estado', {}).get('hierarquia_viaria')
        if hv:
            consolidado['hierarquia_viaria'] = {'dados': hv, 'fonte': fonte}

        # 2) Zonas
        zonas = j['data'].get('estado', {}).get('zonas', {}) or {}
        for z_nome, z_dados in zonas.items():
            if z_nome not in consolidado['zonas']:
                # sigla_canonica eh a sigla limpa (ex: 'ZCA2-A').
                # z_nome pode incluir hierarquia como sufixo (ex: 'ZCA2-A|AP-1') quando ha desambiguacao.
                # v15: UT mais profunda da hierarquia, fallback sigla_canonica, fallback nome
                _hier = z_dados.get('hierarquia') or {}
                sigla_limpa = (_hier.get('UT7') or _hier.get('UT6') or _hier.get('UT5') or
                               _hier.get('UT4') or _hier.get('UT3') or _hier.get('UT2') or _hier.get('UT1') or
                               z_dados.get('sigla_canonica') or z_nome.split('|')[-1])
                consolidado['zonas'][z_nome] = {
                    'sigla': sigla_limpa,
                    'fonte_definicao': fonte,
                    'usos': {},
                    'params_gerais': {},
                    'params_por_uso': {},
                    'variacoes': {},
                    'hierarquia': z_dados.get('hierarquia', {}),
                }

            zc = consolidado['zonas'][z_nome]

            # Usos permitidos: sobrescreve
            usos = z_dados.get('usos_permitidos', {}) or {}
            for u, u_dados in usos.items():
                zc['usos'][u] = {
                    'status': u_dados.get('status'),
                    'condicao': u_dados.get('condicao', ''),
                    'fonte': u_dados.get('fonte') or fonte,
                }

            # Parametros gerais: sobrescreve
            pg = z_dados.get('parametros_gerais', {}) or {}
            for k, v in pg.items():
                if isinstance(v, dict):
                    zc['params_gerais'][k] = {
                        'valor': v.get('valor', v),
                        'fonte': v.get('fonte') or fonte,
                    }
                else:
                    zc['params_gerais'][k] = {'valor': v, 'fonte': fonte}

            # Parametros por uso: sobrescreve
            ppu = z_dados.get('parametros_por_uso', {}) or {}
            for u, params in ppu.items():
                if u not in zc['params_por_uso']:
                    zc['params_por_uso'][u] = {}
                for k, v in (params or {}).items():
                    if isinstance(v, dict):
                        zc['params_por_uso'][u][k] = {
                            'valor': v.get('valor', v),
                            'fonte': v.get('fonte') or fonte,
                        }
                    else:
                        zc['params_por_uso'][u][k] = {'valor': v, 'fonte': fonte}

            # Variacoes
            zc['variacoes'].update(z_dados.get('variacoes', {}) or {})

    return consolidado


def valor_str(v) -> str:
    """Converte valor estruturado em string pra celula."""
    if v is None or v == '':
        return ''
    if isinstance(v, dict):
        # Pode ter 'valor' direto
        if 'valor' in v:
            return valor_str(v['valor'])
        return str(v)
    if isinstance(v, (list, tuple)):
        return '; '.join(valor_str(x) for x in v if x)
    if isinstance(v, bool):
        return 'SIM' if v else 'NÃO'
    return str(v)




# ════════════════════════════════════════════════════════════════════
# PARTE 2 — MAPEAMENTO COLUNAS DA PLANILHA v3.2 (339 cols)
# ════════════════════════════════════════════════════════════════════

# Cada uso ocupa 22 cols (11 params × 2). Bloco comeca em col_inicial.
USO_BLOCO_INICIO = {
    'residencial_unifamiliar':    63,
    'residencial_multifamiliar':  85,
    'residencial_his':           107,
    'comercial':                 129,
    'servicos':                  151,
    'uso_misto':                 173,
    'industrial':                195,
    'institucional':             217,
    'residencial_transitorio_hotel': None,
}

USOS_COL_STATUS = {
    'residencial_unifamiliar':    46,
    'residencial_multifamiliar':  48,
    'residencial_his':            50,
    'comercial':                  52,
    'servicos':                   54,
    'uso_misto':                  56,
    'industrial':                 58,
    'institucional':              60,
    'residencial_transitorio_hotel': None,
}

COL_OBSERVACOES = 62

USO_PARAMS_OFFSET = [
    (0,  'coeficiente_aproveitamento_basico'),
    (2,  'coeficiente_aproveitamento_maximo'),
    (4,  'taxa_ocupacao_basica_pct'),
    (6,  'taxa_ocupacao_maxima_pct'),
    (8,  'gabarito_basico_pavimentos'),
    (10, 'gabarito_max_nao_afastado_pavimentos'),
    (12, 'gabarito_basico_altura_m'),
    (14, 'altura_maxima_absoluta_m'),
    (16, 'recuo_frontal_m'),
    (18, 'recuo_lateral_m'),
    (20, 'recuo_fundos_m'),
]

G1_LOTE = [
    (22, 'area_lote_minimo_m2'),
    (24, 'testada_minima_m'),
    (26, 'area_lote_maximo_m2'),
    (28, 'area_doacao_pct'),
]

G2_GENERAL = [
    (30, 'taxa_permeabilidade_minima_pct'),
    (32, 'quota_ideal_m2_economia'),
    (34, 'gabarito_varia_altitude'),
    (36, 'afastamento_entre_blocos'),
    (38, 'gabarito_max_nao_afastado_pavimentos'),
    (40, 'gabarito_max_nao_afastado_altura_m'),
    (42, 'isencao_outorga_onerosa'),
    (44, 'varia_declividade'),
]

# Mapeamento G11 (cols 328-339): Hierarquia Viaria
# Chaves esperadas no JSON v14 em estado.hierarquia_viaria
G11_HIERARQUIA = [
    (328, 'definida_na_lei'),
    (330, 'hierarquias_existentes'),
    (332, 'vias_arteriais'),
    (334, 'vias_coletoras'),
    (336, 'vias_locais'),
    (338, 'outras_hierarquias'),
]


# ════════════════════════════════════════════════════════════════════
# FUNCOES DE PREENCHIMENTO POR GRUPO
# ════════════════════════════════════════════════════════════════════

def _set_par(ws, linha, col_valor, valor_dict_ou_str, fonte_fallback=''):
    """Preenche (col_valor, col_valor+1) com valor + legislacao."""
    if isinstance(valor_dict_ou_str, dict):
        v = valor_str(valor_dict_ou_str.get('valor'))
        f = valor_dict_ou_str.get('fonte') or fonte_fallback
        # Variacoes por via: acrescenta linhas VIA|nome|lado|valor
        variacoes_via = valor_dict_ou_str.get('variacoes_por_via') or []
        if variacoes_via:
            linhas_via = [f"VIA|{vv.get('via','')}|{vv.get('lado','AMBOS')}|{valor_str(vv.get('valor',''))}"
                          for vv in variacoes_via if vv.get('via')]
            if linhas_via:
                v = (v or '') + '\n' + '\n'.join(linhas_via)
        # Variacoes por hierarquia viaria: acrescenta linhas HIERARQUIA|categoria|valor
        variacoes_hier = valor_dict_ou_str.get('variacoes_por_hierarquia_viaria') or []
        if variacoes_hier:
            linhas_hier = [f"HIERARQUIA|{vh.get('categoria','')}|{valor_str(vh.get('valor',''))}"
                           for vh in variacoes_hier if vh.get('categoria')]
            if linhas_hier:
                v = (v or '') + '\n' + '\n'.join(linhas_hier)
    else:
        v = valor_str(valor_dict_ou_str)
        f = fonte_fallback
    if v not in (None, '', 'None'):
        ws.cell(linha, col_valor).value = v
        if f:
            ws.cell(linha, col_valor + 1).value = f


def preencher_identificacao(ws, linha, zona_dados, municipio, estado, fonte_geral):
    """G1: cols 1-31"""
    ws.cell(linha, 1).value = 'Brasil'
    ws.cell(linha, 2).value = estado
    ws.cell(linha, 3).value = municipio

    # Cols 4-17: UT1-UT7 (v3.3)
    h = zona_dados.get('hierarquia', {}) or {}
    leg_def = zona_dados.get('fonte_definicao', '') or fonte_geral
    for i, ut_key in enumerate(['UT1','UT2','UT3','UT4','UT5','UT6','UT7'], start=1):
        col_val = 2 + i*2  # UT1->4, UT2->6, ..., UT7->16
        ut_val = h.get(ut_key)
        ws.cell(linha, col_val).value = ut_val if ut_val else 'NI'
        ws.cell(linha, col_val + 1).value = leg_def if ut_val else 'NI'
    # Cols 18-21: Zoneamento Ambiental 1 + 2
    za = zona_dados.get('zoneamento_ambiental_sobreposto')
    if za:
        if isinstance(za, list):
            if len(za) > 0: ws.cell(linha, 18).value = str(za[0])
            if len(za) > 1: ws.cell(linha, 20).value = str(za[1])
        else:
            ws.cell(linha, 18).value = str(za)

    # Cols 24-31: Lote (params_gerais) - NI se faltar
    pg = zona_dados.get('params_gerais', {})
    for col, chave in G1_LOTE:
        v = pg.get(chave)
        if v and isinstance(v, dict) and v.get('valor') not in (None, '', 'None'):
            _set_par(ws, linha, col, v, fonte_geral)
        else:
            ws.cell(linha, col).value = 'NI'


def preencher_g2_general(ws, linha, zona_dados, fonte_geral):
    """G2: cols 32-47 — se chave faltar OU valor for vazio, escreve 'NI'."""
    pg = zona_dados.get('params_gerais', {})
    for col, chave in G2_GENERAL:
        v = pg.get(chave) if chave else None
        if v and isinstance(v, dict) and v.get('valor') not in (None, '', 'None'):
            _set_par(ws, linha, col, v, fonte_geral)
        else:
            ws.cell(linha, col).value = 'NI'


def preencher_g3_usos(ws, linha, zona_dados, usos_reconh, fonte_geral):
    """G3: cols 48-66 (status SIM/NAO/NI por uso + Observacoes)"""
    usos_zona = zona_dados.get('usos', {})
    # Acumula condicoes pra preencher COL_OBSERVACOES no final
    observacoes = []
    for uso, col in USOS_COL_STATUS.items():
        if col is None:
            continue  # residencial_transitorio_hotel sem coluna propria
        if uso not in usos_reconh:
            # Lei silenciou completamente sobre este uso → NI
            ws.cell(linha, col).value = 'NI'
            continue
        u_data = usos_zona.get(uso)
        if not u_data:
            # Uso reconhecido pelo municipio mas zona nao lista → assume NÃO
            ws.cell(linha, col).value = 'NÃO'
            continue
        status = (u_data.get('status') or '').upper()
        # SIM, CONDICIONADO → SIM puro (sem condicao na celula)
        if status in ('SIM', 'CONDICIONADO'):
            simplificado = 'SIM'
        elif status in ('NÃO', 'NAO', 'NO'):
            simplificado = 'NÃO'
        else:
            simplificado = status or 'NI'
        ws.cell(linha, col).value = simplificado
        ws.cell(linha, col + 1).value = u_data.get('fonte') or fonte_geral
        # Se houver condicao, acumula pra observacoes
        cond = (u_data.get('condicao') or '').strip()
        if cond and simplificado == 'SIM':
            nome_legivel = uso.replace('_', ' ').title()
            observacoes.append(f"{nome_legivel}: {cond}")
    # col 66: Observacoes - junta as condicoes acumuladas
    if observacoes:
        ws.cell(linha, COL_OBSERVACOES).value = ' | '.join(observacoes)
    else:
        ws.cell(linha, COL_OBSERVACOES).value = 'NI'


def preencher_g4_params_por_uso(ws, linha, zona_dados, usos_reconh, fonte_geral):
    """G4: cols 67-264 (11 params × 2 × 9 usos = 198 cols).
    Regras:
    - Uso silenciado (lei nao cita): tudo 'NI'
    - Uso PROIBIDO (status NÃO): tudo 'NÃO'
    - Uso permitido (SIM/CONDICIONADO): valor real ou 'NI' se faltar dado
    """
    ppu = zona_dados.get('params_por_uso', {})
    usos_zona = zona_dados.get('usos', {})

    for uso, col_inicio in USO_BLOCO_INICIO.items():
        if col_inicio is None:
            continue  # sem coluna na planilha
        # Caso 1: uso silenciado pela lei
        if uso not in usos_reconh:
            for off, _ in USO_PARAMS_OFFSET:
                ws.cell(linha, col_inicio + off).value = 'NI'
            continue
        # Caso 2: uso proibido — preenche 'NOT ALLOWED' em todas as cols
        u_data = usos_zona.get(uso, {})
        status = (u_data.get('status') or '').upper()
        if status in ('NÃO', 'NAO', 'NO'):
            for off, _ in USO_PARAMS_OFFSET:
                ws.cell(linha, col_inicio + off).value = 'NOT ALLOWED'
            continue
        # Caso 3: uso permitido (SIM/CONDICIONADO) - preenche valor real ou NI
        uso_params = ppu.get(uso) or {}
        for off, chave_v14 in USO_PARAMS_OFFSET:
            v = uso_params.get(chave_v14) if chave_v14 else None
            if v and isinstance(v, dict) and v.get('valor') not in (None, '', 'None'):
                _set_par(ws, linha, col_inicio + off, v, fonte_geral)
            else:
                ws.cell(linha, col_inicio + off).value = 'NI'


def preencher_g11_hierarquia_viaria(ws, linha, consolidado):
    """G11: cols 328-339 - somente preenche se v14 forneceu hierarquia_viaria"""
    hv_wrap = consolidado.get('hierarquia_viaria')
    if not hv_wrap:
        return
    hv_data = hv_wrap.get('dados', {}) or {}
    fonte = hv_wrap.get('fonte', '')
    for col, chave in G11_HIERARQUIA:
        val = hv_data.get(chave)
        if val:
            if isinstance(val, list):
                val_str = '; '.join(str(x) for x in val)
            else:
                val_str = str(val)
            ws.cell(linha, col).value = val_str
            ws.cell(linha, col + 1).value = fonte


# ════════════════════════════════════════════════════════════════════
# FUNCAO PRINCIPAL: preencher_planilha
# ════════════════════════════════════════════════════════════════════

def preencher_planilha(template_path: str, consolidado: Dict,
                       municipio: str, estado: str,
                       output_path: str, log_callback=None) -> Dict:
    """Carrega o template, preenche todas as zonas, salva no output_path.
    Retorna {n_zonas, n_jsons_usados, tamanho_bytes, filepath, filename}."""
    from openpyxl import load_workbook

    def _log(m):
        if log_callback: log_callback(m)
        else: logger.info(m)

    _log(f"Carregando template: {template_path}")
    wb = load_workbook(template_path)
    ws = wb.active  # 'Parâmetros Urbanísticos'

    # Conjunto de usos reconhecidos no municipio
    usos_reconh = set()
    for z_dados in consolidado['zonas'].values():
        usos_reconh.update(z_dados.get('usos', {}).keys())
    _log(f"Usos reconhecidos: {len(usos_reconh)}: {sorted(usos_reconh)}")

    # Header: linhas 1-3 reservadas. Comeca em linha 4.
    linha_atual = 5
    n_zonas = 0
    for zona_nome in sorted(consolidado['zonas'].keys()):
        z_dados = consolidado['zonas'][zona_nome]
        fonte_geral = z_dados.get('fonte_definicao', '')
        _log(f"  Zona {zona_nome} → linha {linha_atual}")
        preencher_identificacao(ws, linha_atual, z_dados, municipio, estado, fonte_geral)
        preencher_g2_general(ws, linha_atual, z_dados, fonte_geral)
        preencher_g3_usos(ws, linha_atual, z_dados, usos_reconh, fonte_geral)
        preencher_g4_params_por_uso(ws, linha_atual, z_dados, usos_reconh, fonte_geral)
        preencher_g11_hierarquia_viaria(ws, linha_atual, consolidado)
        linha_atual += 1
        n_zonas += 1

    # Salva
    import os as _os
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    size = _os.path.getsize(output_path)
    _log(f"Salvo: {output_path} ({size} bytes, {n_zonas} zonas)")

    return {
        'n_zonas': n_zonas,
        'tamanho_bytes': size,
        'filepath': output_path,
        'filename': _os.path.basename(output_path),
    }


# ════════════════════════════════════════════════════════════════════
# FUNCAO ORQUESTRADORA: gera_planilha_municipio
# ════════════════════════════════════════════════════════════════════

def gera_planilha_municipio(jsons_ids: List[int], template_path: str,
                            municipio: str, estado: str,
                            get_db_func, log_callback=None) -> Optional[Dict]:
    """Pipeline completo: ids → carregar → consolidar → preencher → salvar."""
    def _log(m):
        if log_callback: log_callback(m)
        else: logger.info(m)

    _log(f"== Preenchedor: {municipio}/{estado} ({len(jsons_ids)} JSON(s)) ==")

    jsons = carregar_jsons(jsons_ids, get_db_func)
    if not jsons:
        _log("ERRO: nenhum JSON valido encontrado")
        return None
    _log(f"Carregados {len(jsons)} JSONs em ordem cronologica:")
    for j in jsons:
        _log(f"  {j['data_publicacao']} → {j['legislacao_label']}")

    # ===== Fase B.8: merge de leis externas =====
    try:
        from modulos.mesclar_leis_externas import mesclar_leis_externas
        jsons_mesclados, log_merges = mesclar_leis_externas(jsons)
        if log_merges:
            _log(f"== Fase B (merge de leis externas): {len(log_merges)} merge(s) realizadas ==")
            for m in log_merges:
                lei_ext = m.get('lei_externa') or {}
                tipo = lei_ext.get('tipo_nome') or lei_ext.get('tipo') or '?'
                numero = lei_ext.get('numero') or '?'
                ano = lei_ext.get('ano') or '?'
                subzonas = m.get('subzonas_adicionadas') or []
                _log(
                    f"  {m.get('lei_pai')} :: '{m.get('zona_pai')}' "
                    f"expandida via {tipo} {numero}/{ano} -> "
                    f"{len(subzonas)} subzona(s): {', '.join(subzonas)}"
                )
            jsons = jsons_mesclados
        else:
            _log("== Fase B: nenhuma lei externa referenciada estava entre os JSONs selecionados (sem merge) ==")
    except Exception as e:
        _log(f"AVISO Fase B (merge): falhou - {str(e)[:200]} (seguindo sem merge)")
        logger.warning(f"Fase B merge falhou: {e}", exc_info=True)

    consolidado = consolidar(jsons)
    _log(f"Consolidado: {len(consolidado['zonas'])} zona(s)")

    # Nome do arquivo
    import os as _os, time as _t
    slug = (municipio or 'mun').replace(' ', '-').replace('/', '-')
    ts = _t.strftime('%d%m%Y_%H%M')
    filename = f"Parametros_{estado}_{slug}_{ts}.xlsx"
    output_dir = f"/var/www/urbanlex/static/planilhas_geradas/{slug}_{estado}"
    output_path = _os.path.join(output_dir, filename)

    meta = preencher_planilha(template_path, consolidado, municipio, estado,
                              output_path, log_callback)
    if not meta:
        return None
    meta['jsons_ids'] = jsons_ids
    meta['template_path'] = template_path
    meta['n_jsons_usados'] = len(jsons)
    return meta
